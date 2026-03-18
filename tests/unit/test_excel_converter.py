"""Unit tests for Excel ↔ UMF bidirectional converter."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from tablespec.excel_converter import (
    ExcelConstants,
    ExcelToUMFConverter,
    ExcelValidator,
    UMFToExcelConverter,
)
from tablespec.models.umf import (
    UMF,
    ForeignKey,
    Nullable,
    OutgoingRelationship,
    Relationships,
    UMFColumn,
    ValidationRule,
    ValidationRules,
)

pytestmark = [pytest.mark.no_spark, pytest.mark.fast]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_minimal_umf(**overrides) -> UMF:
    """Create a minimal valid UMF for testing."""
    defaults = {
        "version": "1.0",
        "table_name": "test_table",
        "canonical_name": "Test Table",
        "columns": [
            UMFColumn(
                name="col_id",
                data_type="INTEGER",
                description="Primary key",
            ),
            UMFColumn(
                name="col_name",
                data_type="VARCHAR",
                length=100,
                description="Name field",
                nullable=Nullable(MD=True, MP=False, ME=True),
                sample_values=["Alice", "Bob"],
            ),
        ],
    }
    defaults.update(overrides)
    return UMF(**defaults)


def _make_rich_umf() -> UMF:
    """Create a UMF with validation rules, relationships, and various column types."""
    return UMF(
        version="2.0",
        table_name="rich_table",
        canonical_name="Rich Table",
        description="A table with many features",
        table_type="provided",
        primary_key=["col_id"],
        aliases=["rich", "Rich Table"],
        columns=[
            UMFColumn(
                name="col_id",
                data_type="INTEGER",
                description="Primary key",
                key_type="primary",
                source="data",
            ),
            UMFColumn(
                name="col_name",
                data_type="VARCHAR",
                length=200,
                description="Full name",
                nullable=Nullable(MD=True, MP=True, ME=False),
                sample_values=["Alice", "Bob"],
                canonical_name="Full Name",
                aliases=["name", "full_name"],
                reporting_requirement="R",
                format="Last, First",
            ),
            UMFColumn(
                name="col_amount",
                data_type="DECIMAL",
                precision=10,
                scale=2,
                description="Dollar amount",
            ),
            UMFColumn(
                name="col_active",
                data_type="BOOLEAN",
                description="Is active flag",
            ),
            UMFColumn(
                name="col_date",
                data_type="DATE",
                description="Created date",
                format="YYYY-MM-DD",
                notes=["Must be after 2020-01-01", "Business calendar only"],
            ),
        ],
        validation_rules=ValidationRules(
            expectations=[
                {
                    "type": "expect_column_values_to_not_be_null",
                    "kwargs": {"column": "col_id"},
                    "meta": {
                        "description": "ID must not be null",
                        "severity": "critical",
                    },
                },
                {
                    "type": "expect_column_values_to_be_in_set",
                    "kwargs": {"column": "col_active", "value_set": [True, False]},
                    "meta": {
                        "description": "Must be boolean",
                        "severity": "warning",
                    },
                },
            ],
        ),
        relationships=Relationships(
            foreign_keys=[
                ForeignKey(
                    column="col_id",
                    references_table="other_table",
                    references_column="id",
                ),
            ],
            outgoing=[
                OutgoingRelationship(
                    source_column="col_id",
                    target_table="other_table",
                    target_column="id",
                    type="foreign_to_primary",
                    confidence=0.95,
                    reasoning="ID match",
                ),
            ],
        ),
    )


@pytest.fixture()
def minimal_umf():
    return _make_minimal_umf()


@pytest.fixture()
def rich_umf():
    return _make_rich_umf()


# ---------------------------------------------------------------------------
# UMFToExcelConverter tests
# ---------------------------------------------------------------------------


class TestUMFToExcelConverter:
    """Test UMF -> Excel export."""

    def test_convert_returns_workbook(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        assert isinstance(wb, openpyxl.Workbook)

    def test_required_sheets_present(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        names = wb.sheetnames
        assert ExcelConstants.SHEET_README in names
        assert ExcelConstants.SHEET_SCHEMA in names
        assert ExcelConstants.SHEET_COLUMNS in names
        assert ExcelConstants.SHEET_METADATA in names

    def test_schema_sheet_values(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        ws = wb[ExcelConstants.SHEET_SCHEMA]
        # Row 2 = table_name, Row 3 = canonical_name
        assert ws["A2"].value == "table_name"
        assert ws["B2"].value == "test_table"
        assert ws["A3"].value == "canonical_name"
        assert ws["B3"].value == "Test Table"

    def test_columns_sheet_headers(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # Check header row
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Canonical Name"
        assert ws["D1"].value == "Data Type"

    def test_columns_sheet_data(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # Row 2 = first column
        assert ws["A2"].value == "col_id"
        assert ws["D2"].value == "INTEGER"
        # Row 3 = second column
        assert ws["A3"].value == "col_name"
        assert ws["D3"].value == "VARCHAR"
        assert ws["E3"].value == 100  # length

    def test_nullable_values_in_columns_sheet(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # col_name (row 3) has nullable MD=True, MP=False, ME=True
        # Context keys are sorted alphabetically: MD, ME, MP
        assert ws["H3"].value is True  # MD
        assert ws["I3"].value is True  # ME
        assert ws["J3"].value is False  # MP

    def test_sample_values_in_columns_sheet(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # col_name has sample_values ["Alice", "Bob"]
        assert ws["L3"].value == "Alice, Bob"

    def test_validation_sheet_created_when_rules_exist(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        assert ExcelConstants.SHEET_VALIDATION in wb.sheetnames

    def test_validation_sheet_not_created_when_no_rules(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        assert ExcelConstants.SHEET_VALIDATION not in wb.sheetnames

    def test_relationships_sheet_created_when_fks_exist(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        assert ExcelConstants.SHEET_RELATIONSHIPS in wb.sheetnames

    def test_relationships_sheet_not_created_when_no_fks(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        assert ExcelConstants.SHEET_RELATIONSHIPS not in wb.sheetnames

    def test_save_to_file(self, minimal_umf, tmp_path):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        out = tmp_path / "test.xlsx"
        wb.save(out)
        assert out.exists()
        # Re-open to ensure file is valid
        wb2 = openpyxl.load_workbook(out)
        assert ExcelConstants.SHEET_SCHEMA in wb2.sheetnames

    def test_description_in_schema(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        ws = wb[ExcelConstants.SHEET_SCHEMA]
        # Find description row
        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == "description":
                assert row[1].value == "A table with many features"
                found = True
                break
        assert found, "description field not found in Schema sheet"

    def test_decimal_column_precision_scale(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # Find col_amount row
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            if row_cells[0].value == "col_amount":
                assert row_cells[3].value == "DECIMAL"  # data_type
                assert row_cells[4].value == ""  # length (not applicable)
                assert row_cells[5].value == 10  # precision
                assert row_cells[6].value == 2  # scale
                return
        pytest.fail("col_amount not found in Columns sheet")

    def test_notes_exported_as_newline_string(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # Find col_date row
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            if row_cells[0].value == "col_date":
                notes_val = row_cells[17].value  # R column = index 17
                assert "Must be after 2020-01-01" in notes_val
                assert "Business calendar only" in notes_val
                return
        pytest.fail("col_date not found in Columns sheet")

    def test_format_exported(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            if row_cells[0].value == "col_name":
                assert row_cells[16].value == "Last, First"  # Q column = index 16
                return
        pytest.fail("col_name not found in Columns sheet")

    def test_primary_key_in_schema(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        ws = wb[ExcelConstants.SHEET_SCHEMA]
        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == "primary_key":
                assert row[1].value == "col_id"
                found = True
                break
        assert found, "primary_key not found in Schema sheet"

    def test_aliases_in_columns(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            if row_cells[0].value == "col_name":
                aliases_val = row_cells[2].value  # C column = aliases
                assert "name" in aliases_val
                assert "full_name" in aliases_val
                return
        pytest.fail("col_name not found")


# ---------------------------------------------------------------------------
# ExcelValidator tests
# ---------------------------------------------------------------------------


class TestExcelValidator:
    """Test Excel workbook validation."""

    def test_valid_workbook_passes(self, minimal_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(minimal_umf)
        validator = ExcelValidator()
        assert validator.validate_workbook(wb) is True

    def test_missing_schema_sheet_fails(self):
        wb = openpyxl.Workbook()
        # Only default sheet, no Schema or Columns
        validator = ExcelValidator()
        assert validator.validate_workbook(wb) is False
        assert any("Schema" in e for e in validator.errors)

    def test_missing_columns_sheet_fails(self):
        wb = openpyxl.Workbook()
        wb.create_sheet("Schema")
        ws = wb["Schema"]
        ws["A2"] = "table_name"
        ws["B2"] = "test"
        ws["A3"] = "canonical_name"
        ws["B3"] = "Test"
        validator = ExcelValidator()
        assert validator.validate_workbook(wb) is False
        assert any("Columns" in e for e in validator.errors)

    def test_rich_workbook_passes(self, rich_umf):
        converter = UMFToExcelConverter()
        wb = converter.convert(rich_umf)
        validator = ExcelValidator()
        result = validator.validate_workbook(wb)
        # If there are errors, print them for debugging
        if not result:
            pytest.fail(f"Validation failed: {validator.errors}")


# ---------------------------------------------------------------------------
# ExcelToUMFConverter tests (individual extraction methods)
# ---------------------------------------------------------------------------


class TestExcelToUMFConverterExtractSchema:
    """Test _extract_schema method."""

    def test_extract_schema_from_exported_workbook(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        schema = importer._extract_schema(wb)
        assert schema["table_name"] == "rich_table"
        assert schema["canonical_name"] == "Rich Table"
        assert schema["description"] == "A table with many features"
        assert schema["version"] == "2.0"
        assert schema.get("table_type") == "provided"

    def test_extract_schema_aliases(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        schema = importer._extract_schema(wb)
        # Aliases should be parsed as list; "Rich Table" filtered out by exporter since it matches canonical_name
        if "aliases" in schema:
            assert isinstance(schema["aliases"], list)

    def test_extract_schema_primary_key(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        schema = importer._extract_schema(wb)
        assert "primary_key" in schema
        assert "col_id" in schema["primary_key"]


class TestExcelToUMFConverterExtractColumns:
    """Test _extract_columns method."""

    def test_extract_columns_from_exported_workbook(self, minimal_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(minimal_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        assert len(columns) == 2
        assert columns[0]["name"] == "col_id"
        assert columns[1]["name"] == "col_name"

    def test_extract_columns_data_types(self, minimal_umf):
        """Data types are normalized to Spark types by the importer."""
        exporter = UMFToExcelConverter()
        wb = exporter.convert(minimal_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        # INTEGER maps to IntegerType, VARCHAR maps to StringType
        assert columns[0]["data_type"] == "IntegerType"
        assert columns[1]["data_type"] == "StringType"

    def test_extract_columns_length(self, minimal_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(minimal_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        assert columns[1].get("length") == 100

    def test_extract_columns_nullable(self, minimal_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(minimal_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        # col_name has nullable MD=True, MP=False, ME=True
        nullable = columns[1].get("nullable")
        assert nullable is not None
        assert nullable["MD"] is True
        assert nullable["MP"] is False
        assert nullable["ME"] is True

    def test_extract_columns_sample_values(self, minimal_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(minimal_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        sv = columns[1].get("sample_values")
        assert sv is not None
        assert "Alice" in sv
        assert "Bob" in sv

    def test_extract_columns_description(self, minimal_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(minimal_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        assert columns[0].get("description") == "Primary key"

    def test_extract_empty_columns_sheet(self):
        """A workbook with no data rows in Columns sheet returns empty list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Columns"
        ws["A1"] = "Name"
        ws["D1"] = "Data Type"
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        assert columns == []

    def test_extract_columns_decimal_precision_scale(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(wb)
        amount_col = next(c for c in columns if c["name"] == "col_amount")
        assert amount_col.get("precision") == 10
        assert amount_col.get("scale") == 2


class TestExcelToUMFConverterExtractValidation:
    """Test _extract_validation method."""

    def test_extract_validation_rules(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        validation, review_notes = importer._extract_validation(wb)
        assert validation is not None
        expectations = validation.get("expectations", [])
        assert len(expectations) >= 2

    def test_extract_validation_rule_types(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        validation, _ = importer._extract_validation(wb)
        expectations = validation["expectations"]
        types = [e["type"] for e in expectations]
        assert "expect_column_values_to_not_be_null" in types
        assert "expect_column_values_to_be_in_set" in types

    def test_extract_validation_returns_empty_when_no_sheet(self):
        wb = openpyxl.Workbook()
        importer = ExcelToUMFConverter()
        validation, notes = importer._extract_validation(wb)
        assert validation is None
        assert notes == {}


class TestExcelToUMFConverterExtractRelationships:
    """Test _extract_relationships method."""

    def test_extract_relationships(self, rich_umf):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        importer = ExcelToUMFConverter()
        relationships = importer._extract_relationships(wb)
        assert relationships is not None
        fks = relationships.get("foreign_keys", [])
        assert len(fks) >= 1
        assert fks[0]["column"] == "col_id"
        # The exporter writes target_table/target_column, importer reads as references_table/references_column
        assert fks[0]["references_table"] == "other_table"
        assert fks[0]["references_column"] == "id"


class TestExcelToUMFConverterConvert:
    """Test full convert() method via file on disk."""

    def test_convert_file_not_found(self, tmp_path):
        importer = ExcelToUMFConverter()
        with pytest.raises(FileNotFoundError):
            importer.convert(tmp_path / "nonexistent.xlsx")

    def test_convert_invalid_workbook(self, tmp_path):
        """A workbook without required sheets should raise ValueError."""
        wb = openpyxl.Workbook()
        out = tmp_path / "bad.xlsx"
        wb.save(out)
        importer = ExcelToUMFConverter()
        with pytest.raises(ValueError, match="Invalid Excel workbook"):
            importer.convert(out)


# ---------------------------------------------------------------------------
# Round-trip tests (export then manually verify sheet content)
# ---------------------------------------------------------------------------


class TestExcelRoundTrip:
    """Test that exporting to Excel and reading back preserves data at sheet level."""

    def test_round_trip_via_file(self, rich_umf, tmp_path):
        """Export UMF -> Excel file -> reload workbook -> verify sheet data."""
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        out = tmp_path / "round_trip.xlsx"
        wb.save(out)

        wb2 = openpyxl.load_workbook(out)
        ws = wb2[ExcelConstants.SHEET_COLUMNS]

        # Verify columns are present
        col_names = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            name = row[0].value
            if name:
                col_names.append(name)
        assert "col_id" in col_names
        assert "col_name" in col_names
        assert "col_amount" in col_names
        assert "col_active" in col_names
        assert "col_date" in col_names

    def test_round_trip_schema_preserved(self, rich_umf, tmp_path):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        out = tmp_path / "rt_schema.xlsx"
        wb.save(out)

        importer = ExcelToUMFConverter()
        schema = importer._extract_schema(openpyxl.load_workbook(out))
        assert schema["table_name"] == "rich_table"
        assert schema["canonical_name"] == "Rich Table"
        assert schema["version"] == "2.0"

    def test_round_trip_column_count_preserved(self, rich_umf, tmp_path):
        exporter = UMFToExcelConverter()
        wb = exporter.convert(rich_umf)
        out = tmp_path / "rt_cols.xlsx"
        wb.save(out)

        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(openpyxl.load_workbook(out))
        assert len(columns) == len(rich_umf.columns)

    def test_round_trip_nullable_preserved(self, tmp_path):
        umf = _make_minimal_umf()
        exporter = UMFToExcelConverter()
        wb = exporter.convert(umf)
        out = tmp_path / "rt_nullable.xlsx"
        wb.save(out)

        importer = ExcelToUMFConverter()
        columns = importer._extract_columns(openpyxl.load_workbook(out))
        col_name_data = next(c for c in columns if c["name"] == "col_name")
        nullable = col_name_data["nullable"]
        assert nullable["MD"] is True
        assert nullable["MP"] is False
        assert nullable["ME"] is True


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    """Test edge cases and boundary conditions."""

    def test_single_column_umf(self):
        umf = UMF(
            version="1.0",
            table_name="one_col",
            columns=[UMFColumn(name="only_col", data_type="VARCHAR", length=50)],
        )
        converter = UMFToExcelConverter()
        wb = converter.convert(umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        assert ws["A2"].value == "only_col"
        # Row 3 should be empty
        assert ws["A3"].value is None

    def test_column_with_no_optional_fields(self):
        umf = UMF(
            version="1.0",
            table_name="bare",
            columns=[UMFColumn(name="bare_col", data_type="INTEGER")],
        )
        converter = UMFToExcelConverter()
        wb = converter.convert(umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        assert ws["A2"].value == "bare_col"
        assert ws["D2"].value == "INTEGER"
        # No nullable, no sample values
        assert ws["H2"].value is None
        assert ws["L2"].value == ""

    def test_all_data_types_export(self):
        """Verify all UMF data types can be exported to Excel."""
        cols = []
        for dtype in ["VARCHAR", "INTEGER", "DECIMAL", "DATE", "DATETIME", "BOOLEAN", "TEXT", "CHAR", "FLOAT"]:
            extra = {}
            if dtype == "VARCHAR":
                extra = {"length": 50}
            elif dtype == "DECIMAL":
                extra = {"precision": 10, "scale": 2}
            elif dtype == "CHAR":
                extra = {"length": 1}
            cols.append(
                UMFColumn(name=f"col_{dtype.lower()}", data_type=dtype, **extra)
            )
        umf = UMF(version="1.0", table_name="all_types", columns=cols)
        converter = UMFToExcelConverter()
        wb = converter.convert(umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        exported_types = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value:
                exported_types.append(row[3].value)
        assert len(exported_types) == len(cols)

    def test_empty_description(self):
        umf = UMF(
            version="1.0",
            table_name="no_desc",
            columns=[UMFColumn(name="c", data_type="INTEGER")],
        )
        converter = UMFToExcelConverter()
        wb = converter.convert(umf)
        ws = wb[ExcelConstants.SHEET_COLUMNS]
        # Description cell should be empty string, not None
        assert ws["K2"].value in (None, "")
