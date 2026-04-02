"""Excel ↔ UMF bidirectional converter with strict validation.

Enables round-trip conversion between Excel workbooks and UMF schemas.
Excel format is designed for non-technical domain experts with:
- Data validation (dropdowns, constraints)
- Helper columns (validation status, error messages)
- VBA macros (for guidance and validation)
- Clear instructions and examples
"""

import contextlib
import json
import logging
from pathlib import Path
from typing import Any, ClassVar

import openpyxl
import openpyxl.cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from tablespec.models import UMF, Nullable
from tablespec.naming import position_sort_key


logger = logging.getLogger(__name__)


class ExcelConstants:
    """Constants for Excel workbook structure."""

    # Sheet names
    SHEET_README = "README"
    SHEET_SCHEMA = "Schema"
    SHEET_COLUMNS = "Columns"
    SHEET_VALIDATION = "Validation Rules"
    SHEET_RELATIONSHIPS = "Relationships"
    SHEET_FILE_FORMAT = "File Format"
    SHEET_METADATA = "Metadata"
    SHEET_INSTRUCTIONS = "_Instructions"
    SHEET_VALIDATION_HIDDEN = "_Validation"

    # Data type options (PySpark type names)
    DATA_TYPES: ClassVar[list[str]] = [
        "StringType",
        "IntegerType",
        "LongType",
        "ShortType",
        "ByteType",
        "DecimalType",
        "FloatType",
        "DoubleType",
        "DateType",
        "TimestampType",
        "BooleanType",
    ]

    # Key type options
    KEY_TYPES: ClassVar[list[str]] = [
        "primary",
        "unique",
        "foreign_many_to_one",
        "foreign_one_to_one",
    ]

    # Source options
    SOURCES: ClassVar[list[str]] = ["data", "filename", "metadata"]

    # Reporting requirement options
    REPORTING_REQUIREMENTS: ClassVar[list[str]] = ["R", "O", "S"]

    # Table type options
    TABLE_TYPES: ClassVar[list[str]] = ["provided", "generated", "lookup"]

    # Default nullable context keys when no per-context nullable dicts exist in UMF data.
    # Used as fallback for backward compatibility with existing Excel templates.
    DEFAULT_NULLABLE_CONTEXTS: ClassVar[list[str]] = ["MD", "MP", "ME"]

    # Domain types - fallback list if registry is unavailable
    _DEFAULT_DOMAIN_TYPES: ClassVar[list[str]] = [
        "member_id",
        "claim_id",
        "provider_id",
        "npi",
        "date_of_birth",
        "date_of_service",
        "medical_record_number",
        "us_state_code",
        "us_county_code",
        "zip_code",
        "email",
        "phone_number",
        "currency_amount",
        "account_number",
    ]

    @staticmethod
    def _get_domain_type_list() -> list[str]:
        """Get domain types from registry, falling back to hardcoded defaults."""
        try:
            from tablespec.inference.domain_types import DomainTypeRegistry

            registry = DomainTypeRegistry()
            return sorted(registry.list_domain_types())
        except Exception:
            return sorted(ExcelConstants._DEFAULT_DOMAIN_TYPES)

    @property
    def DOMAIN_TYPES(self) -> list[str]:  # noqa: N802
        """Domain types loaded dynamically from registry."""
        return self._get_domain_type_list()

    # Rule types (Great Expectations expectation types - without "expect_" prefix)
    # Based on expectation_categories.json and expectation_parameters.json
    RULE_TYPES: ClassVar[list[str]] = [
        # Baseline (generated from UMF)
        "column_to_exist",
        "column_values_to_be_of_type",
        "column_values_to_not_be_null",
        "column_value_lengths_to_be_between",
        "column_values_to_match_strftime_format",
        "table_column_count_to_equal",
        "table_columns_to_match_ordered_list",
        "table_columns_to_match_set",
        # LLM Table-level
        "compound_columns_to_be_unique",
        "column_pair_values_a_to_be_greater_than_b",
        "column_pair_values_to_be_equal",
        "table_row_count_to_be_between",
        "table_row_count_to_equal_other_table",
        "select_column_values_to_be_unique_within_record",
        # LLM Column-level
        "column_values_to_be_in_set",
        "column_values_to_match_regex",
        "column_values_to_match_regex_list",
        "column_values_to_be_unique",
        "column_values_to_be_between",
        "column_distinct_values_to_be_in_set",
        "column_distinct_values_to_contain_set",
        "column_most_common_value_to_be_in_set",
        "column_values_to_not_be_in_set",
        "column_values_to_not_match_regex",
        "column_values_to_not_match_regex_list",
        "column_value_lengths_to_equal",
        "column_values_to_be_increasing",
        "column_values_to_be_decreasing",
        "column_values_to_be_json_parseable",
        "column_values_to_match_json_schema",
        "column_values_to_cast_to_type",
        "column_values_to_be_null",
        # Statistical (profiling)
        "column_min_to_be_between",
        "column_max_to_be_between",
        "column_mean_to_be_between",
        "column_median_to_be_between",
        "column_stdev_to_be_between",
        "column_sum_to_be_between",
        "column_kl_divergence_to_be_less_than",
        "column_proportion_of_non_null_values_to_be_between",
        "column_proportion_of_unique_values_to_be_between",
        "column_unique_value_count_to_be_between",
        # Advanced
        "column_values_to_be_dateutil_parseable",
        "column_values_to_be_in_type_list",
        # Pending fallback
        "validation_rule_pending_implementation",
    ]

    # Severities (GX standard severity levels + skip for disabled rules)
    SEVERITIES: ClassVar[list[str]] = ["critical", "warning", "info", "skip"]

    # Colors
    COLOR_ERROR = "FF0000"
    COLOR_WARNING = "FFFF00"
    COLOR_SUCCESS = "00B050"
    COLOR_HEADER = "4472C4"
    COLOR_HELPER = "E7E6E6"
    COLOR_HIGHLIGHT = "FFF2CC"

    # Font sizes
    FONT_SIZE_TITLE = 15
    FONT_SIZE_HEADER = 13
    FONT_SIZE_DEFAULT = 13


class ExcelValidator:
    """Validates Excel workbook for UMF compatibility."""

    def __init__(self) -> None:
        """Initialize validator."""
        self.errors = []
        self.warnings = []

    def validate_workbook(self, workbook: openpyxl.Workbook) -> bool:
        """Validate entire workbook structure and content.

        Args:
            workbook: Excel workbook to validate

        Returns:
            True if valid, False otherwise

        """
        self.errors.clear()
        self.warnings.clear()

        # Check required sheets
        required_sheets = [
            ExcelConstants.SHEET_SCHEMA,
            ExcelConstants.SHEET_COLUMNS,
        ]

        sheet_names = workbook.sheetnames
        for sheet in required_sheets:
            if sheet not in sheet_names:
                self.errors.append(f"Missing required sheet: {sheet}")

        if self.errors:
            return False

        # Validate schema sheet
        self._validate_schema_sheet(workbook)

        # Validate columns sheet
        self._validate_columns_sheet(workbook)

        # Validate relationships if present
        if ExcelConstants.SHEET_RELATIONSHIPS in sheet_names:
            self._validate_relationships_sheet(workbook)

        return len(self.errors) == 0

    def _validate_schema_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Validate schema sheet contents."""
        ws = workbook[ExcelConstants.SHEET_SCHEMA]

        # Check required fields
        required_fields = ["table_name", "canonical_name"]
        for field in required_fields:
            value = self._get_schema_value(ws, field)
            if not value:
                self.errors.append(f"Schema: Missing required field '{field}'")

    def _validate_columns_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Validate columns sheet contents."""
        from tablespec.type_mappings import map_to_gx_spark_type

        ws = workbook[ExcelConstants.SHEET_COLUMNS]

        # Check that columns exist
        data_rows = self._get_data_rows(ws)
        if not data_rows:
            self.errors.append("Columns: At least one column must be defined")
            return

        # Validate each column
        # Column layout: Name(0), Canonical Name(1), Aliases(2), Data Type(3), Length(4), Precision(5), Scale(6)
        for row_num, row in enumerate(data_rows, start=2):  # Start at 2 (skip header)
            col_name = self._get_cell_value(row, 0)
            data_type = self._get_cell_value(row, 3)  # Data Type is now column D (index 3)

            if not col_name:
                self.errors.append(f"Columns row {row_num}: Missing column name")

            normalized_type = None
            if not data_type:
                self.errors.append(f"Columns row {row_num}: Missing data type")
            else:
                # Accept both SparkSQL and SQL-style types (will be normalized during import)
                normalized_type = map_to_gx_spark_type(data_type)
                if normalized_type not in ExcelConstants.DATA_TYPES:
                    self.errors.append(f"Columns row {row_num}: Invalid data type '{data_type}'")

            # Validate type-specific requirements (use normalized type)
            if data_type and normalized_type == "StringType":
                length = self._get_cell_value(row, 4)  # Length is now column E (index 4)
                if not length:
                    self.warnings.append(
                        f"Columns row {row_num}: StringType columns should specify length"
                    )

            elif data_type and normalized_type == "DecimalType":
                precision = self._get_cell_value(row, 5)  # Precision is now column F (index 5)
                scale = self._get_cell_value(row, 6)  # Scale is now column G (index 6)
                if not precision or not scale:
                    self.errors.append(
                        f"Columns row {row_num}: DECIMAL/DecimalType requires precision and scale"
                    )

    def _validate_relationships_sheet(self, workbook: openpyxl.Workbook) -> None:
        """Validate relationships sheet contents."""
        ws = workbook[ExcelConstants.SHEET_RELATIONSHIPS]
        data_rows = self._get_data_rows(ws)

        columns_ws = workbook[ExcelConstants.SHEET_COLUMNS]
        column_names = self._get_column_names(columns_ws)

        for row_num, row in enumerate(data_rows, start=2):
            source_col = self._get_cell_value(row, 0)

            if source_col and source_col not in column_names:
                self.errors.append(
                    f"Relationships row {row_num}: Source column '{source_col}' not found in Columns sheet"
                )

    def _get_schema_value(self, ws: Worksheet, field: str) -> Any:
        """Get value from schema sheet."""
        field_row_map = {
            "table_name": 2,
            "canonical_name": 3,
            "description": 5,
        }

        row = field_row_map.get(field)
        if row:
            return ws[f"B{row}"].value
        return None

    def _get_data_rows(self, ws: Worksheet) -> list:
        """Get all data rows from sheet (skip header, skip empty rows)."""
        rows = []
        for _row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Check if row has any data
            has_data = any(cell.value for cell in row)
            if has_data:
                rows.append(row)
        return rows

    def _get_cell_value(self, row: tuple, col_index: int) -> Any:
        """Safely get cell value from row."""
        if col_index < len(row):
            return row[col_index].value
        return None

    def _get_column_names(self, ws: Worksheet) -> list[str]:
        """Get all column names from Columns sheet."""
        names = []
        for row in self._get_data_rows(ws):
            name = self._get_cell_value(row, 0)
            if name:
                names.append(name)
        return names


class UMFToExcelConverter:
    """Convert UMF to Excel workbook."""

    def __init__(self) -> None:
        """Initialize converter."""
        self.workbook: openpyxl.Workbook | None = None
        self.constants = ExcelConstants()

    @staticmethod
    def _extract_context_keys(umf: UMF) -> list[str]:
        """Extract nullable context keys from UMF columns.

        Scans all columns' nullable fields to collect the union of context keys.
        Returns them in stable sorted order. Falls back to the default contexts
        (MD, MP, ME) when no per-context nullable dicts are found.
        """
        keys: set[str] = set()
        for col in umf.columns:
            if col.nullable is None:
                continue
            if isinstance(col.nullable, Nullable):
                keys.update(col.nullable.model_dump(exclude_none=True).keys())
            elif isinstance(col.nullable, dict):
                keys.update(col.nullable.keys())
        if not keys:
            return list(ExcelConstants.DEFAULT_NULLABLE_CONTEXTS)
        return sorted(keys)

    def _get_default_font(self) -> Font:
        """Get default font for data cells."""
        return Font(size=self.constants.FONT_SIZE_DEFAULT)

    def _get_header_font(self) -> Font:
        """Get font for header rows."""
        return Font(bold=True, color="FFFFFF", size=self.constants.FONT_SIZE_HEADER)

    def _get_title_font(self) -> Font:
        """Get font for titles."""
        return Font(bold=True, size=self.constants.FONT_SIZE_TITLE)

    def _apply_font_to_cell(self, cell: Any, font: Font) -> None:
        """Apply font to a cell."""
        cell.font = font

    def convert(self, umf: UMF) -> openpyxl.Workbook:
        """Convert UMF to Excel workbook.

        Args:
            umf: UMF model to convert

        Returns:
            Excel workbook

        Raises:
            ValueError: If UMF is invalid

        """
        self.workbook = openpyxl.Workbook()
        if self.workbook.active is not None:
            self.workbook.remove(self.workbook.active)  # Remove default sheet

        # Create hidden instructions sheet first (contains dropdown data for validations)
        self._create_instructions_sheet()

        # Create sheets
        self._create_readme_sheet(umf)
        self._create_schema_sheet(umf)
        self._create_columns_sheet(umf)
        self._create_survivorship_sheet(umf)

        if umf.validation_rules:
            self._create_validation_sheet(umf)

        if umf.relationships:
            self._create_relationships_sheet(umf)

        if umf.file_format:
            self._create_file_format_sheet(umf)

        self._create_metadata_sheet(umf)

        return self.workbook

    def _create_readme_sheet(self, umf: UMF) -> None:
        """Create README sheet with instructions and status."""
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_README, 0)

        # Title
        ws["A1"] = "Universal Metadata Format (UMF) Excel Editor"
        self._apply_font_to_cell(ws["A1"], self._get_title_font())

        # Status
        ws["A3"] = "Validation Status:"
        ws["B3"] = "✓ Valid"  # Update based on validation
        ws["B3"].font = Font(color="00B050", bold=True, size=self.constants.FONT_SIZE_DEFAULT)
        self._apply_font_to_cell(ws["A3"], self._get_default_font())

        # Instructions
        row = 5
        ws[f"A{row}"] = "Instructions:"
        self._apply_font_to_cell(
            ws[f"A{row}"], Font(bold=True, size=self.constants.FONT_SIZE_DEFAULT)
        )

        instructions = [
            "1. Fill in Schema sheet with table metadata",
            "2. Add columns in Columns sheet (name, type, description, etc.)",
            "3. Define validation rules (optional)",
            "4. Define relationships (optional)",
            "5. Configure file format (optional)",
            "6. Export to UMF format when complete",
        ]

        row += 1
        for instruction in instructions:
            ws[f"A{row}"] = instruction
            self._apply_font_to_cell(ws[f"A{row}"], self._get_default_font())
            row += 1

        # Export section
        row += 2
        ws[f"A{row}"] = "Export Options:"
        self._apply_font_to_cell(
            ws[f"A{row}"], Font(bold=True, size=self.constants.FONT_SIZE_DEFAULT)
        )
        ws[f"A{row + 1}"] = "Run: tablespec convert-from-excel sheet.xlsx table.umf.yaml"
        self._apply_font_to_cell(ws[f"A{row + 1}"], self._get_default_font())
        ws[f"A{row + 2}"] = "Or: tablespec convert-from-excel sheet.xlsx table.umf.json"
        self._apply_font_to_cell(ws[f"A{row + 2}"], self._get_default_font())

        # Adjust columns
        ws.column_dimensions["A"].width = 80

    def _create_schema_sheet(self, umf: UMF) -> None:
        """Create Schema sheet with table-level metadata."""
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_SCHEMA)

        # Header
        ws["A1"] = "Field"
        ws["B1"] = "Value"
        self._style_header_row(ws, 1)

        # Filter out canonical_name from table aliases (case-insensitive)
        filtered_table_aliases = []
        if umf.aliases:
            canonical_lower = umf.canonical_name.lower() if umf.canonical_name else ""
            filtered_table_aliases = [a for a in umf.aliases if a.lower() != canonical_lower]

        # Data
        data = [
            ("table_name", umf.table_name),
            ("canonical_name", umf.canonical_name),
            ("version", umf.version),
            ("description", umf.description or ""),
            ("aliases", ", ".join(filtered_table_aliases) if filtered_table_aliases else ""),
            ("source_file", umf.source_file or ""),
            ("source_sheet_name", umf.source_sheet_name or ""),
            ("table_type", umf.table_type or ""),
            ("primary_key", ", ".join(umf.primary_key) if umf.primary_key else ""),
        ]

        row = 2
        for field, value in data:
            ws[f"A{row}"] = field
            self._apply_font_to_cell(ws[f"A{row}"], self._get_default_font())
            ws[f"B{row}"] = value
            self._apply_font_to_cell(ws[f"B{row}"], self._get_default_font())
            row += 1

        # Add data validation for table_type (row 9, column B)
        dv_table_type = DataValidation(
            type="list",
            formula1=f'"{",".join(self.constants.TABLE_TYPES)}"',
            allow_blank=True,
        )
        dv_table_type.add("B9")  # table_type is at row 9
        ws.add_data_validation(dv_table_type)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    def _create_columns_sheet(self, umf: UMF) -> None:
        """Create Columns sheet with column definitions.

        Nullable context columns are derived dynamically from the UMF data.
        For example, if columns use ``nullable: {MD: false, MP: true}``, the
        sheet will contain "Nullable MD" and "Nullable MP" headers. When no
        per-context nullable dicts exist, the default contexts (MD, MP, ME)
        are used for backward compatibility.
        """
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_COLUMNS)

        # Derive nullable context keys from UMF data
        context_keys = self._extract_context_keys(umf)
        num_contexts = len(context_keys)

        # Fixed columns before nullable: indices 0-6
        # A=Name, B=Canonical Name, C=Aliases, D=Data Type, E=Length, F=Precision, G=Scale
        NULLABLE_START = 7  # First nullable column index (0-based)

        # Headers
        headers = [
            "Name",
            "Canonical Name",
            "Aliases",
            "Data Type",
            "Length",
            "Precision",
            "Scale",
        ]
        for key in context_keys:
            headers.append(f"Nullable {key}")
        # Columns after nullable
        post_nullable_headers = [
            "Description",
            "Sample Values",
            "Source",
            "Key Type",
            "Domain Type",
            "Reporting Req",
            "Format",
            "Notes",
            "_Validation",
        ]
        headers.extend(post_nullable_headers)

        self._add_header_row(ws, headers)
        self._add_data_validation_to_columns(ws, num_contexts=num_contexts)

        # Helper: column letter from 0-based index
        def col_letter(idx: int) -> str:
            return get_column_letter(idx + 1)

        # Offsets for post-nullable fields (0-based indices)
        desc_idx = NULLABLE_START + num_contexts
        sample_idx = desc_idx + 1
        source_idx = desc_idx + 2
        key_type_idx = desc_idx + 3
        domain_type_idx = desc_idx + 4
        reporting_idx = desc_idx + 5
        format_idx = desc_idx + 6
        notes_idx = desc_idx + 7

        # Data
        row = 2
        for col in umf.columns:
            default_font = self._get_default_font()
            ws[f"A{row}"] = col.name
            self._apply_font_to_cell(ws[f"A{row}"], default_font)
            ws[f"B{row}"] = col.canonical_name or ""
            self._apply_font_to_cell(ws[f"B{row}"], default_font)

            # Filter out canonical_name and table_name from aliases (case-insensitive)
            filtered_aliases = []
            if col.aliases:
                canonical_lower = col.canonical_name.lower() if col.canonical_name else ""
                table_lower = umf.table_name.lower() if umf.table_name else ""
                filtered_aliases = [
                    a
                    for a in col.aliases
                    if a.lower() != canonical_lower and a.lower() != table_lower
                ]
            ws[f"C{row}"] = ", ".join(filtered_aliases) if filtered_aliases else ""
            self._apply_font_to_cell(ws[f"C{row}"], default_font)
            ws[f"D{row}"] = col.data_type
            self._apply_font_to_cell(ws[f"D{row}"], default_font)
            ws[f"E{row}"] = col.length or ""
            self._apply_font_to_cell(ws[f"E{row}"], default_font)
            ws[f"F{row}"] = col.precision or ""
            self._apply_font_to_cell(ws[f"F{row}"], default_font)
            ws[f"G{row}"] = col.scale or ""
            self._apply_font_to_cell(ws[f"G{row}"], default_font)

            # Nullable - write dynamic context columns
            if col.nullable:
                nullable_data: dict[str, bool] = {}
                if isinstance(col.nullable, Nullable):
                    nullable_data = col.nullable.model_dump(exclude_none=True)
                elif isinstance(col.nullable, dict):
                    nullable_data = col.nullable

                for i, key in enumerate(context_keys):
                    cell_ref = f"{col_letter(NULLABLE_START + i)}{row}"
                    ws[cell_ref] = nullable_data.get(key, False)
                    self._apply_font_to_cell(ws[cell_ref], default_font)
            else:
                for i in range(num_contexts):
                    cell_ref = f"{col_letter(NULLABLE_START + i)}{row}"
                    self._apply_font_to_cell(ws[cell_ref], default_font)

            ws[f"{col_letter(desc_idx)}{row}"] = col.description or ""
            self._apply_font_to_cell(ws[f"{col_letter(desc_idx)}{row}"], default_font)
            ws[f"{col_letter(sample_idx)}{row}"] = (
                ", ".join(col.sample_values) if col.sample_values else ""
            )
            self._apply_font_to_cell(ws[f"{col_letter(sample_idx)}{row}"], default_font)
            ws[f"{col_letter(source_idx)}{row}"] = col.source or "data"
            self._apply_font_to_cell(ws[f"{col_letter(source_idx)}{row}"], default_font)
            ws[f"{col_letter(key_type_idx)}{row}"] = col.key_type or ""
            self._apply_font_to_cell(ws[f"{col_letter(key_type_idx)}{row}"], default_font)
            ws[f"{col_letter(domain_type_idx)}{row}"] = col.domain_type or ""
            self._apply_font_to_cell(ws[f"{col_letter(domain_type_idx)}{row}"], default_font)
            ws[f"{col_letter(reporting_idx)}{row}"] = col.reporting_requirement or ""
            self._apply_font_to_cell(ws[f"{col_letter(reporting_idx)}{row}"], default_font)

            # Format
            ws[f"{col_letter(format_idx)}{row}"] = col.format or ""
            self._apply_font_to_cell(ws[f"{col_letter(format_idx)}{row}"], default_font)

            # Notes (list[str] -> newline-separated string)
            if col.notes:
                ws[f"{col_letter(notes_idx)}{row}"] = "\n".join(col.notes)
            else:
                ws[f"{col_letter(notes_idx)}{row}"] = ""
            self._apply_font_to_cell(ws[f"{col_letter(notes_idx)}{row}"], default_font)

            row += 1

        # Adjust column widths
        pre_nullable_widths = [15, 18, 20, 12, 10, 11, 8]
        nullable_widths = [12] * num_contexts
        post_nullable_widths = [20, 20, 12, 18, 15, 12, 15, 30, 15]
        all_widths = pre_nullable_widths + nullable_widths + post_nullable_widths
        for i, width in enumerate(all_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_survivorship_sheet(self, umf: UMF) -> None:
        """Create Survivorship sheet with hierarchical column-organized format.

        Structure:
        - Level 1 rows: Column-level summary (position, reporting requirement, provenance, strategy, explanation)
        - Level 2 rows: Individual candidates (priority, source, reason)
        - Excel grouping applied to collapse/expand candidate details
        - All columns included, sorted by position field
        """
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet("Survivorship")

        # Headers
        headers = [
            "Level",
            "Column Name",
            "Position",
            "Reporting Requirement",
            "Provenance",
            "Strategy",
            "Priority",
            "Source",
            "Explanation/Reason",
        ]

        # Add header row with styling
        header_font = Font(name="Arial", size=11, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Data rows
        default_font = Font(name="Arial", size=11)
        row = 2

        # Sort columns by position field using shared utility
        sorted_columns = sorted(
            umf.columns, key=lambda col: position_sort_key(col.position, umf.columns.index(col))
        )

        for col in sorted_columns:
            # Level 1: Column-level summary row
            ws.cell(row, 1).value = 1  # Level
            ws.cell(row, 2).value = col.name  # Column Name

            # Position
            ws.cell(row, 3).value = col.position or ""

            # Reporting Requirement
            ws.cell(row, 4).value = col.reporting_requirement or ""

            # Provenance
            provenance = col.provenance_policy or "survivorship" if col.derivation else ""
            ws.cell(row, 5).value = provenance

            # Strategy
            strategy = ""
            if col.derivation and col.derivation.survivorship:
                strategy = col.derivation.survivorship.strategy or ""
            ws.cell(row, 6).value = strategy

            # Priority and Source are blank for Level 1
            ws.cell(row, 7).value = ""
            ws.cell(row, 8).value = ""

            # Explanation with rich text formatting for bold description
            explanation = ""
            if (
                col.derivation
                and col.derivation.survivorship
                and col.derivation.survivorship.explanation
            ):
                explanation = col.derivation.survivorship.explanation

            # Parse and format explanation with bold description if present
            if explanation and explanation.startswith("Requirement: "):
                # Split into: "Requirement: ", description, separator, main explanation
                parts = explanation.split("\n\n---\n\n", 1)
                if len(parts) == 2:
                    requirement_and_desc = parts[0]  # "Requirement: {description}"
                    main_explanation = parts[1]

                    # Extract description (everything after "Requirement: ")
                    desc_text = requirement_and_desc.replace("Requirement: ", "", 1)

                    # Create rich text with bold description
                    rich_text = CellRichText(
                        "Requirement: ",
                        TextBlock(InlineFont(b=True), desc_text),
                        "\n\n---\n\n",
                        main_explanation,
                    )
                    ws.cell(row, 9).value = rich_text
                else:
                    # Fallback if separator not found
                    ws.cell(row, 9).value = explanation
            else:
                ws.cell(row, 9).value = explanation

            ws.cell(row, 9).alignment = Alignment(wrap_text=True, vertical="top")

            # Apply font and top alignment to Level 1 row
            for col_idx in range(1, 10):
                cell = ws.cell(row, col_idx)
                self._apply_font_to_cell(cell, default_font)
                if col_idx != 9:  # Column 9 already has alignment set above
                    cell.alignment = Alignment(vertical="top")

            row += 1

            # Level 2: Candidate detail rows (if any)
            first_candidate_row = None
            if col.derivation and col.derivation.candidates:
                first_candidate_row = row

                for candidate in sorted(col.derivation.candidates, key=lambda c: c.priority):
                    ws.cell(row, 1).value = 2  # Level
                    ws.cell(row, 2).value = ""  # Column Name (blank for Level 2)
                    ws.cell(row, 3).value = ""  # Position (blank for Level 2)
                    ws.cell(row, 4).value = ""  # Reporting Requirement (blank for Level 2)
                    ws.cell(row, 5).value = ""  # Provenance (blank for Level 2)
                    ws.cell(row, 6).value = ""  # Strategy (blank for Level 2)
                    ws.cell(row, 7).value = candidate.priority  # Priority

                    # Build source column representation
                    source_parts = [candidate.table]

                    # Add table_instance as alias if present (e.g., "outreach_list_pcp (pcp_assigned)")
                    if candidate.table_instance:
                        source_parts[0] = f"{candidate.table} ({candidate.table_instance})"

                    # Add column or expression
                    if candidate.expression:
                        source_parts.append(candidate.expression)
                    else:
                        source_parts.append(candidate.column or "")

                    ws.cell(row, 8).value = ".".join(source_parts)  # Source

                    # Build reason with optional join_filter
                    reason_parts = []
                    if candidate.reason:
                        reason_parts.append(candidate.reason)
                    if candidate.join_filter:
                        reason_parts.append(f"Filter: {candidate.join_filter}")

                    ws.cell(row, 9).value = (
                        "\n\n".join(reason_parts) if reason_parts else ""
                    )  # Reason
                    ws.cell(row, 9).alignment = Alignment(wrap_text=True, vertical="top")

                    # Apply font and top alignment to Level 2 row
                    for col_idx in range(1, 10):
                        cell = ws.cell(row, col_idx)
                        self._apply_font_to_cell(cell, default_font)
                        if col_idx != 9:  # Column 9 already has alignment set above
                            cell.alignment = Alignment(vertical="top")

                    row += 1

            # Level 2: Default value row (if any)
            if (
                col.derivation
                and col.derivation.survivorship
                and col.derivation.survivorship.default_value is not None
            ):
                if first_candidate_row is None:
                    first_candidate_row = row

                ws.cell(row, 1).value = 2  # Level
                ws.cell(row, 2).value = ""  # Column Name (blank for Level 2)
                ws.cell(row, 3).value = ""  # Position (blank for Level 2)
                ws.cell(row, 4).value = ""  # Reporting Requirement (blank for Level 2)
                ws.cell(row, 5).value = ""  # Provenance (blank for Level 2)
                ws.cell(row, 6).value = ""  # Strategy (blank for Level 2)
                ws.cell(row, 7).value = "default"  # Priority (string)
                ws.cell(row, 8).value = str(
                    col.derivation.survivorship.default_value
                )  # Source (default value)
                ws.cell(row, 9).value = (
                    col.derivation.survivorship.default_condition or ""
                )  # Reason (default condition)
                ws.cell(row, 9).alignment = Alignment(wrap_text=True, vertical="top")

                # Apply font and top alignment to Level 2 row
                for col_idx in range(1, 10):
                    cell = ws.cell(row, col_idx)
                    self._apply_font_to_cell(cell, default_font)
                    if col_idx != 9:  # Column 9 already has alignment set above
                        cell.alignment = Alignment(vertical="top")

                row += 1

            # Group Level 2 rows under Level 1 (collapsible)
            if first_candidate_row is not None and row > first_candidate_row:
                ws.row_dimensions.group(first_candidate_row, row - 1, outline_level=1, hidden=False)

        # Set column widths
        ws.column_dimensions["A"].width = 8  # Level
        ws.column_dimensions["B"].width = 25  # Column Name
        ws.column_dimensions["C"].width = 12  # Position
        ws.column_dimensions["D"].width = 15  # Reporting Requirement
        ws.column_dimensions["E"].width = 18  # Provenance
        ws.column_dimensions["F"].width = 20  # Strategy
        ws.column_dimensions["G"].width = 10  # Priority
        ws.column_dimensions["H"].width = 30  # Source
        ws.column_dimensions["I"].width = 100  # Explanation/Reason (very wide for readability)

    def _prepare_validation_expectations_with_index(
        self, expectations: list[dict[str, Any]]
    ) -> tuple[list[tuple[str, str, str, int, dict[str, Any]]], list[str]]:
        """Prepare expectations with index assignment and sorting.

        Groups expectations by (column, rule_type) and assigns indices.
        Severity is NOT part of the grouping key - it's a property that can change.
        Preserves existing rule_index from metadata if available (for round-trip edits).

        Returns:
            Tuple of (sorted_expectations, parameter_names_sorted)
            where sorted_expectations is list of (column, severity, rule_type, index, expectation)
            and parameter_names_sorted is alphabetically sorted list of all parameter names

        """
        from collections import defaultdict

        # Build a map of (column, rule_type) -> list of expectations
        # NOTE: severity is NOT in the key - only column and rule_type
        groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)

        # Collect all unique parameter names (excluding "column" which is handled separately)
        param_names: set[str] = set()

        for exp in expectations:
            # Extract rule type (remove "expect_" prefix)
            exp_type = exp.get("type", "")
            rule_type = exp_type.removeprefix("expect_")

            # Extract column name, or "-" for table-level
            column_name = exp.get("kwargs", {}).get("column", "-")

            # Collect parameter names from kwargs (exclude "column")
            kwargs = exp.get("kwargs", {})
            for param_name in kwargs:
                if param_name != "column":
                    param_names.add(param_name)

            # Group by (column, rule_type) - severity is NOT part of the key
            key = (column_name, rule_type)
            groups[key].append(exp)

        # Build result list with indices
        # Sort so column-specific rules come first (alphabetically), then cross-column rules (column="-")
        def sort_key(item):
            column_name, rule_type = item[0]
            # Put "-" (cross-column) last by using (1, "-"), all others first with (0, column_name)
            priority = 1 if column_name == "-" else 0
            return (priority, column_name, rule_type)

        result = []
        for (column_name, rule_type), exps in sorted(groups.items(), key=sort_key):
            # Separate expectations with existing indices from new ones
            with_index = []
            without_index = []

            for exp in exps:
                rule_index = exp.get("meta", {}).get("rule_index")
                if rule_index is not None:
                    with_index.append((rule_index, exp))
                else:
                    without_index.append(exp)

            # Sort by existing index, then assign new indices to new expectations
            with_index.sort(key=lambda x: x[0])

            # Find the next available index
            existing_indices = {idx for idx, _ in with_index}
            next_index = 0
            for rule_index, exp in with_index:
                severity = exp.get("meta", {}).get("severity", "error")
                result.append((column_name, severity, rule_type, rule_index, exp))

            # Assign new indices to expectations without rule_index
            for exp in without_index:
                while next_index in existing_indices:
                    next_index += 1
                severity = exp.get("meta", {}).get("severity", "error")
                result.append((column_name, severity, rule_type, next_index, exp))
                existing_indices.add(next_index)
                next_index += 1

        # Sort parameter names alphabetically
        param_names_sorted = sorted(param_names)

        return result, param_names_sorted

    def _create_validation_sheet(self, umf: UMF) -> None:
        """Create Validation Rules sheet with key | value | key | value format.

        Fixed columns:
        - Review Notes: Empty column for manual review tracking
        - Column: Field name or "-" for table-level
        - Severity: critical, warning, info
        - Rule Type: Expectation type (without "expect_" prefix)
        - Index: 0-based index for multiple rules of same type on same column
        - Description: Human-readable description

        Dynamic columns (alternating key/value pairs):
        - Key1 | Value1 | Key2 | Value2 | Key3 | Value3 ...

        Example:
        Review Notes | Column | Severity | Rule Type | Index | Description | column_list | [col1,col2] | value_set | [A,B,C]

        """
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_VALIDATION)

        # Build headers: fixed columns + space for parameters
        headers = [
            "Review Notes",
            "Column",
            "Severity",
            "Rule Type",
            "Index",
            "Description",
            "Generated From",
            "Rule ID",
            "LOB",
            "Reason Unmappable",
            "Suggested Implementation",
            "Domain Type",
        ]
        self._add_header_row(ws, headers)

        if umf.validation_rules and umf.validation_rules.expectations:
            expectations_with_index, _ = self._prepare_validation_expectations_with_index(
                umf.validation_rules.expectations
            )

            default_font = self._get_default_font()
            max_col = 12  # Start after Domain Type column (L)

            row = 2
            for column_name, severity, rule_type, index, exp in expectations_with_index:
                # Review Notes - empty column for manual entry
                ws[f"A{row}"] = ""
                self._apply_font_to_cell(ws[f"A{row}"], default_font)

                # Fixed columns (shifted right by 1)
                ws[f"B{row}"] = column_name
                self._apply_font_to_cell(ws[f"B{row}"], default_font)
                # Map legacy "error" to "critical" for backward compatibility
                ws[f"C{row}"] = "critical" if severity == "error" else severity
                self._apply_font_to_cell(ws[f"C{row}"], default_font)
                ws[f"D{row}"] = rule_type
                self._apply_font_to_cell(ws[f"D{row}"], default_font)
                ws[f"E{row}"] = index
                self._apply_font_to_cell(ws[f"E{row}"], default_font)
                ws[f"F{row}"] = exp.get("meta", {}).get("description", "")
                self._apply_font_to_cell(ws[f"F{row}"], default_font)
                ws[f"G{row}"] = exp.get("meta", {}).get("generated_from", "")
                self._apply_font_to_cell(ws[f"G{row}"], default_font)
                ws[f"H{row}"] = exp.get("meta", {}).get("rule_id", "")
                self._apply_font_to_cell(ws[f"H{row}"], default_font)

                # LOB (Line of Business) - can be a list
                lob_value = exp.get("meta", {}).get("lob", "")
                if isinstance(lob_value, list):
                    lob_value = ", ".join(lob_value)
                ws[f"I{row}"] = lob_value
                self._apply_font_to_cell(ws[f"I{row}"], default_font)

                ws[f"J{row}"] = exp.get("meta", {}).get("reason_unmappable", "")
                self._apply_font_to_cell(ws[f"J{row}"], default_font)
                ws[f"K{row}"] = exp.get("meta", {}).get("suggested_implementation", "")
                self._apply_font_to_cell(ws[f"K{row}"], default_font)
                ws[f"L{row}"] = exp.get("meta", {}).get("domain_type", "")
                self._apply_font_to_cell(ws[f"L{row}"], default_font)

                # Dynamic parameter columns: key | value | key | value ...
                # (now starting at column 13 instead of 12)
                kwargs = exp.get("kwargs", {})
                col_offset = 0
                for param_name in sorted(kwargs.keys()):
                    if param_name == "column":
                        continue
                    param_value = kwargs[param_name]

                    # Key column
                    col_letter = get_column_letter(13 + col_offset)
                    ws[f"{col_letter}{row}"] = param_name
                    self._apply_font_to_cell(ws[f"{col_letter}{row}"], default_font)

                    # Value column
                    col_letter = get_column_letter(13 + col_offset + 1)
                    if isinstance(param_value, (list, dict)):
                        ws[f"{col_letter}{row}"] = json.dumps(param_value)
                    else:
                        ws[f"{col_letter}{row}"] = param_value
                    self._apply_font_to_cell(ws[f"{col_letter}{row}"], default_font)

                    col_offset += 2
                    max_col = max(max_col, 13 + col_offset - 1)

                row += 1

            # Set column widths
            ws.column_dimensions["A"].width = 40  # Review Notes
            ws.column_dimensions["B"].width = 20  # Column
            ws.column_dimensions["C"].width = 12  # Severity
            ws.column_dimensions["D"].width = 20  # Rule Type
            ws.column_dimensions["E"].width = 8  # Index
            ws.column_dimensions["F"].width = 40  # Description
            ws.column_dimensions["G"].width = 15  # Generated From
            ws.column_dimensions["H"].width = 40  # Rule ID
            ws.column_dimensions["I"].width = 15  # LOB
            ws.column_dimensions["J"].width = 40  # Reason Unmappable
            ws.column_dimensions["K"].width = 40  # Suggested Implementation
            ws.column_dimensions["L"].width = 20  # Domain Type

            # Key/value columns (start at 13)
            for col_num in range(13, max_col + 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 20
        else:
            # No expectations, set minimal widths
            ws.column_dimensions["A"].width = 40  # Review Notes
            ws.column_dimensions["B"].width = 20  # Column
            ws.column_dimensions["C"].width = 12  # Severity
            ws.column_dimensions["D"].width = 20  # Rule Type
            ws.column_dimensions["E"].width = 8  # Index
            ws.column_dimensions["F"].width = 40  # Description
            ws.column_dimensions["G"].width = 15  # Generated From
            ws.column_dimensions["H"].width = 40  # Rule ID

        # Add data validation dropdowns using range references from hidden sheet
        # Severity dropdown (column C) - references _Instructions!M2:M4
        dv_severity = DataValidation(
            type="list",
            formula1=f"'{self.constants.SHEET_INSTRUCTIONS}'!$M$2:$M${2 + len(self.constants.SEVERITIES) - 1}",
            allow_blank=True,
        )
        dv_severity.add("C2:C1000")
        ws.add_data_validation(dv_severity)

        # Rule Type dropdown (column D) - references _Instructions!J2:J46
        dv_rule_type = DataValidation(
            type="list",
            formula1=f"'{self.constants.SHEET_INSTRUCTIONS}'!$J$2:$J${2 + len(self.constants.RULE_TYPES) - 1}",
            allow_blank=True,
        )
        dv_rule_type.add("D2:D1000")
        ws.add_data_validation(dv_rule_type)

    def _create_relationships_sheet(self, umf: UMF) -> None:
        """Create Relationships sheet.

        Columns:
        - Source Column: Column in this table
        - References Table: Table being referenced
        - References Column: Column in referenced table
        - Confidence: Confidence score (0-1)
        - Type: Relationship type (e.g., 'foreign_key')
        - Cardinality: Cardinality notation (e.g., 'N:1', '1:1')
        - Domain Context: Business/domain context for the relationship
        """
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_RELATIONSHIPS)

        headers = [
            "Source Column",
            "References Table",
            "References Column",
            "Confidence",
            "Type",
            "Cardinality",
            "Domain Context",
            "Detection Method",
        ]
        self._add_header_row(ws, headers)

        if umf.relationships and umf.relationships.outgoing:
            default_font = self._get_default_font()
            row = 2
            for fk in umf.relationships.outgoing:
                ws[f"A{row}"] = fk.source_column
                self._apply_font_to_cell(ws[f"A{row}"], default_font)
                ws[f"B{row}"] = fk.target_table
                self._apply_font_to_cell(ws[f"B{row}"], default_font)
                ws[f"C{row}"] = fk.target_column
                self._apply_font_to_cell(ws[f"C{row}"], default_font)
                ws[f"D{row}"] = fk.confidence or ""
                self._apply_font_to_cell(ws[f"D{row}"], default_font)
                ws[f"E{row}"] = fk.type or ""
                self._apply_font_to_cell(ws[f"E{row}"], default_font)
                ws[f"F{row}"] = fk.cardinality.notation if fk.cardinality else ""
                self._apply_font_to_cell(ws[f"F{row}"], default_font)
                ws[f"G{row}"] = fk.reasoning or ""
                self._apply_font_to_cell(ws[f"G{row}"], default_font)
                ws[f"H{row}"] = "ai"
                self._apply_font_to_cell(ws[f"H{row}"], default_font)
                row += 1

        for i, width in enumerate([20, 20, 20, 12, 15, 12, 40, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_file_format_sheet(self, umf: UMF) -> None:
        """Create File Format sheet."""
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_FILE_FORMAT)

        # Headers
        ws["A1"] = "Field"
        ws["B1"] = "Value"
        self._style_header_row(ws, 1)

        # Data
        ff = umf.file_format

        # Build filename_pattern as JSON if present
        filename_pattern_value = ""
        filename_format_value = ""
        if ff and ff.filename_pattern:
            filename_pattern_value = json.dumps(
                {
                    "regex": ff.filename_pattern.regex,
                    "captures": ff.filename_pattern.captures,
                }
            )

            # Generate human-readable filename_format from pattern
            # Convert regex capture groups to {field_name} placeholders
            import re

            format_str = ff.filename_pattern.regex

            # Replace capturing groups with field names in order
            sorted_captures = sorted(ff.filename_pattern.captures.items(), key=lambda x: int(x[0]))

            for _group_num, field_name in sorted_captures:
                # Replace the first capturing group with the field name
                format_str = re.sub(r"\([^)]+\)", f"{{{field_name}}}", format_str, count=1)

            # Clean up regex syntax to make it human-readable
            format_str = format_str.replace("\\d", "").replace("\\", "")
            format_str = format_str.replace("^", "").replace("$", "")
            format_str = format_str.replace("{3,10}", "").replace("{2}", "").replace("{8}", "")
            # Remove optional group markers like (?:...)? but keep the content
            format_str = re.sub(r"\(\?:", "", format_str)
            format_str = re.sub(r"\)\?", "", format_str)
            # Clean up any remaining parentheses
            format_str = format_str.replace("(", "").replace(")", "")

            filename_format_value = format_str

        data = [
            ("delimiter", ff.delimiter if ff else ""),
            ("encoding", ff.encoding if ff else "utf-8"),
            ("header", ff.header if ff else True),
            ("quote_char", ff.quote_char if ff else ""),
            ("escape_char", ff.escape_char if ff else ""),
            ("null_value", ff.null_value if ff else ""),
            ("skip_rows", ff.skip_rows if ff else 0),
            ("comment_char", ff.comment_char if ff else ""),
            ("filename_format", filename_format_value),
            ("filename_pattern", filename_pattern_value),
        ]

        row = 2
        for field, value in data:
            ws[f"A{row}"] = field
            self._apply_font_to_cell(ws[f"A{row}"], self._get_default_font())
            ws[f"B{row}"] = value
            self._apply_font_to_cell(ws[f"B{row}"], self._get_default_font())
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60

    def _create_metadata_sheet(self, umf: UMF) -> None:
        """Create Metadata sheet."""
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_METADATA)

        ws["A1"] = "Field"
        ws["B1"] = "Value"
        self._style_header_row(ws, 1)

        metadata = umf.metadata.model_dump(exclude_none=True) if umf.metadata else {}
        row = 2
        for key, value in metadata.items():
            ws[f"A{row}"] = key
            self._apply_font_to_cell(ws[f"A{row}"], self._get_default_font())
            ws[f"B{row}"] = str(value)
            self._apply_font_to_cell(ws[f"B{row}"], self._get_default_font())
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    def _create_instructions_sheet(self) -> None:
        """Create instructions sheet with data dictionary for data validation references.

        Note: This sheet is kept visible (not hidden) because Excel may reject
        data validation formulas that reference hidden sheets, causing validation
        rules to be silently removed when opening the file.
        """
        if self.workbook is None:
            msg = "Workbook not initialized"
            raise RuntimeError(msg)
        ws = self.workbook.create_sheet(self.constants.SHEET_INSTRUCTIONS)
        # DON'T hide this sheet - Excel doesn't like validation references to hidden sheets
        # ws.sheet_state = "hidden"

        # Data types
        ws["A1"] = "Data Types"
        row = 2
        for dtype in self.constants.DATA_TYPES:
            ws[f"A{row}"] = dtype
            row += 1

        # Key types
        ws["D1"] = "Key Types"
        row = 2
        for ktype in self.constants.KEY_TYPES:
            ws[f"D{row}"] = ktype
            row += 1

        # Sources
        ws["G1"] = "Sources"
        row = 2
        for source in self.constants.SOURCES:
            ws[f"G{row}"] = source
            row += 1

        # Rule types (for validation sheet dropdown)
        ws["J1"] = "Rule Types"
        row = 2
        for rule_type in self.constants.RULE_TYPES:
            ws[f"J{row}"] = rule_type
            row += 1

        # Severities (for validation sheet dropdown)
        ws["M1"] = "Severities"
        row = 2
        for severity in self.constants.SEVERITIES:
            ws[f"M{row}"] = severity
            row += 1

    def _add_header_row(self, ws: Worksheet, headers: list[str]) -> None:
        """Add header row to sheet."""
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            # Skip setting value if cell is a MergedCell (read-only)
            if not isinstance(cell, openpyxl.cell.MergedCell):
                cell.value = header
        self._style_header_row(ws, 1)

    def _style_header_row(self, ws: Worksheet, row_num: int) -> None:
        """Style header row."""
        header_font = self._get_header_font()
        for cell in ws[row_num]:
            self._apply_font_to_cell(cell, header_font)
            cell.fill = PatternFill(
                start_color=self.constants.COLOR_HEADER,
                end_color=self.constants.COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _add_data_validation_to_columns(
        self, ws: Worksheet, *, num_contexts: int = 3
    ) -> None:
        """Add data validation to columns sheet.

        Args:
            ws: The worksheet to add validation to.
            num_contexts: Number of nullable context columns (determines
                where subsequent columns start).
        """
        NULLABLE_START = 7  # 0-based index of first nullable column

        # Data type validation (column D)
        dv_dtype = DataValidation(
            type="list",
            formula1=f'"{",".join(self.constants.DATA_TYPES)}"',
            allow_blank=True,
        )
        dv_dtype.error = "Please select a valid data type"
        dv_dtype.errorTitle = "Invalid Data Type"
        dv_dtype.add("D2:D1000")
        ws.add_data_validation(dv_dtype)

        # Nullable checkboxes - dynamic number of columns starting at NULLABLE_START
        dv_nullable = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
        )
        for i in range(num_contexts):
            col_letter = get_column_letter(NULLABLE_START + i + 1)
            dv_nullable.add(f"{col_letter}2:{col_letter}1000")
        ws.add_data_validation(dv_nullable)

        # Post-nullable column positions (0-based)
        source_col = get_column_letter(NULLABLE_START + num_contexts + 2 + 1)  # +2 for Desc, Sample
        key_type_col = get_column_letter(NULLABLE_START + num_contexts + 3 + 1)
        domain_type_col = get_column_letter(NULLABLE_START + num_contexts + 4 + 1)

        # Source validation
        dv_source = DataValidation(
            type="list",
            formula1=f'"{",".join(self.constants.SOURCES)}"',
            allow_blank=True,
        )
        dv_source.add(f"{source_col}2:{source_col}1000")
        ws.add_data_validation(dv_source)

        # Key type validation
        dv_keytype = DataValidation(
            type="list",
            formula1=f'"{",".join(self.constants.KEY_TYPES)}"',
            allow_blank=True,
        )
        dv_keytype.add(f"{key_type_col}2:{key_type_col}1000")
        ws.add_data_validation(dv_keytype)

        # Domain type validation
        dv_domain = DataValidation(
            type="list",
            formula1=f'"{",".join(self.constants.DOMAIN_TYPES)}"',
            allow_blank=True,
        )
        dv_domain.add(f"{domain_type_col}2:{domain_type_col}1000")
        ws.add_data_validation(dv_domain)

        # Reporting requirement validation
        reporting_col = get_column_letter(NULLABLE_START + num_contexts + 5 + 1)
        dv_reporting = DataValidation(
            type="list",
            formula1=f'"{",".join(self.constants.REPORTING_REQUIREMENTS)}"',
            allow_blank=True,
        )
        dv_reporting.add(f"{reporting_col}2:{reporting_col}1000")
        ws.add_data_validation(dv_reporting)


class ExcelToUMFConverter:
    """Convert Excel workbook to UMF."""

    def __init__(self) -> None:
        """Initialize converter."""
        self.validator = ExcelValidator()
        self.constants = ExcelConstants()

    def convert(self, workbook_path: Path) -> tuple[UMF, dict[str, str | None]]:
        """Convert Excel workbook to UMF.

        Args:
            workbook_path: Path to Excel file

        Returns:
            Tuple of (UMF model, review_notes_map)
            review_notes_map: Map of validation rule index to review note text

        Raises:
            FileNotFoundError: If file not found
            ValueError: If workbook is invalid
            ValidationError: If UMF validation fails

        """
        if not workbook_path.exists():
            msg = f"Excel file not found: {workbook_path}"
            raise FileNotFoundError(msg)

        workbook = openpyxl.load_workbook(workbook_path)

        # Validate workbook
        if not self.validator.validate_workbook(workbook):
            errors_str = "\n".join(self.validator.errors)
            msg = f"Invalid Excel workbook:\n{errors_str}"
            raise ValueError(msg)

        # Extract data
        schema_data = self._extract_schema(workbook)
        columns_data = self._extract_columns(workbook)

        # Build UMF
        umf_dict = {
            **schema_data,
            "columns": columns_data,
        }

        # Add optional sections and extract review notes
        review_notes = {}
        if self.constants.SHEET_VALIDATION in workbook.sheetnames:
            validation_data, validation_notes = self._extract_validation(workbook)
            if validation_data:
                umf_dict["validation_rules"] = validation_data
            review_notes.update(validation_notes)

        if self.constants.SHEET_RELATIONSHIPS in workbook.sheetnames:
            relationships_data = self._extract_relationships(workbook)
            if relationships_data:
                umf_dict["relationships"] = relationships_data

        if self.constants.SHEET_FILE_FORMAT in workbook.sheetnames:
            file_format_data = self._extract_file_format(workbook)
            if file_format_data:
                umf_dict["file_format"] = file_format_data

        if self.constants.SHEET_METADATA in workbook.sheetnames:
            metadata_data = self._extract_metadata(workbook)
            if metadata_data:
                umf_dict["metadata"] = metadata_data

        # Create UMF model (validates schema)
        umf = UMF.model_validate(umf_dict)
        return umf, review_notes

    def _extract_schema(self, workbook: openpyxl.Workbook) -> dict:
        """Extract schema data from Schema sheet."""
        ws = workbook[self.constants.SHEET_SCHEMA]

        schema = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            field = row[0].value
            value = row[1].value

            if field and value:
                if field == "aliases" and isinstance(value, str):
                    schema["aliases"] = [a.strip() for a in value.split(",")]
                elif field == "primary_key" and isinstance(value, str):
                    schema["primary_key"] = [p.strip() for p in value.split(",")]
                elif field in (
                    "version",
                    "table_name",
                    "canonical_name",
                    "description",
                    "table_type",
                    "source_file",
                    "source_sheet_name",
                ):
                    schema[field] = value

        return schema

    @staticmethod
    def _detect_nullable_columns(ws: Worksheet) -> list[tuple[int, str]]:
        """Detect nullable context columns from the header row.

        Returns a list of (column_index, context_key) tuples for headers
        matching the pattern "Nullable <KEY>".
        """
        nullable_cols: list[tuple[int, str]] = []
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for idx, header in enumerate(header_row):
            if header and isinstance(header, str) and header.startswith("Nullable "):
                context_key = header[len("Nullable "):]
                nullable_cols.append((idx, context_key))
        return nullable_cols

    @staticmethod
    def _build_header_index(
        header_row: tuple[Any, ...],
        nullable_indices: set[int],
    ) -> dict[str, int]:
        """Build a mapping from lowercase header names to column indices.

        Skips nullable columns (already handled separately).
        """
        header_map: dict[str, int] = {}
        for idx, header in enumerate(header_row):
            if idx not in nullable_indices and header and isinstance(header, str):
                header_map[header.lower().strip()] = idx
        return header_map

    def _extract_columns(self, workbook: openpyxl.Workbook) -> list[dict]:
        """Extract column definitions from Columns sheet.

        Detects nullable context columns dynamically from the header row
        (any column named "Nullable <KEY>"). Post-nullable fields are
        located by their header names, so the reader handles any number
        of context columns.
        """
        from tablespec.type_mappings import map_to_gx_spark_type

        ws = workbook[self.constants.SHEET_COLUMNS]

        # Detect nullable columns and build header index
        nullable_cols = self._detect_nullable_columns(ws)
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        nullable_indices = {idx for idx, _ in nullable_cols}
        header_map = self._build_header_index(header_row, nullable_indices)

        def _get(row: tuple, idx: int | None) -> Any:
            """Safely get a cell value by index."""
            if idx is None or idx >= len(row):
                return None
            return row[idx].value

        # Resolve post-nullable column indices by header name
        desc_idx = header_map.get("description")
        sample_idx = header_map.get("sample values")
        source_idx = header_map.get("source")
        key_type_idx = header_map.get("key type")
        domain_type_idx = header_map.get("domain type")
        reporting_idx = header_map.get("reporting req")
        derived_from_idx = header_map.get("derived from")
        derivation_mapping_idx = header_map.get("derivation mapping")
        derivation_expression_idx = header_map.get("derivation expression")
        format_idx = header_map.get("format")
        notes_idx = header_map.get("notes")

        columns = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            name = row[0].value
            if not name:  # Skip empty rows
                continue

            col_dict: dict[str, Any] = {
                "name": name,
            }

            # Canonical name (column B)
            if len(row) > 1 and row[1].value:
                col_dict["canonical_name"] = row[1].value

            # Aliases (column C)
            if len(row) > 2 and row[2].value:
                aliases_value = row[2].value
                if isinstance(aliases_value, str):
                    col_dict["aliases"] = [a.strip() for a in aliases_value.split(",")]
                else:
                    col_dict["aliases"] = [a.strip() for a in str(aliases_value).split(",")]

            # Normalize data type
            raw_type = row[3].value if len(row) > 3 else None
            if raw_type and isinstance(raw_type, str):
                normalized_type = map_to_gx_spark_type(raw_type)
            else:
                normalized_type = "StringType"
            col_dict["data_type"] = normalized_type

            # Optional fields
            if len(row) > 4 and row[4].value:  # Length
                length_value = row[4].value
                if isinstance(length_value, (int, float, str)):
                    col_dict["length"] = int(length_value)

            if len(row) > 5 and row[5].value:  # Precision
                precision_value = row[5].value
                if isinstance(precision_value, (int, float, str)):
                    col_dict["precision"] = int(precision_value)

            if len(row) > 6 and row[6].value:  # Scale
                scale_value = row[6].value
                if isinstance(scale_value, (int, float, str)):
                    col_dict["scale"] = int(scale_value)

            # Nullable - read dynamic context columns from detected headers
            if nullable_cols:
                nullable_dict: dict[str, bool] = {}
                any_set = False
                for col_idx, context_key in nullable_cols:
                    val = row[col_idx].value if len(row) > col_idx else None
                    if val is not None:
                        nullable_dict[context_key] = bool(val)
                        any_set = True
                    else:
                        nullable_dict[context_key] = False
                if any_set:
                    col_dict["nullable"] = nullable_dict

            # Post-nullable fields resolved by header name
            desc_val = _get(row, desc_idx)
            if desc_val:
                col_dict["description"] = desc_val

            sample_val = _get(row, sample_idx)
            if sample_val:
                if isinstance(sample_val, str):
                    col_dict["sample_values"] = [s.strip() for s in sample_val.split(",")]
                else:
                    col_dict["sample_values"] = [
                        s.strip() for s in str(sample_val).split(",")
                    ]

            source_val = _get(row, source_idx)
            if source_val:
                col_dict["source"] = source_val

            key_type_val = _get(row, key_type_idx)
            if key_type_val:
                col_dict["key_type"] = key_type_val

            domain_type_val = _get(row, domain_type_idx)
            if domain_type_val:
                col_dict["domain_type"] = domain_type_val

            reporting_val = _get(row, reporting_idx)
            if reporting_val:
                col_dict["reporting_requirement"] = reporting_val

            derived_val = _get(row, derived_from_idx)
            if derived_val:
                col_dict["derived_from"] = derived_val

            # Derivation Mapping (JSON format)
            derivation_map_val = _get(row, derivation_mapping_idx)
            if derivation_map_val:
                if isinstance(derivation_map_val, str):
                    with contextlib.suppress(json.JSONDecodeError, TypeError):
                        col_dict["derivation_mapping"] = json.loads(derivation_map_val)

            # Derivation Expression
            derivation_expr_val = _get(row, derivation_expression_idx)
            if derivation_expr_val:
                col_dict["derivation_expression"] = derivation_expr_val

            # Format
            format_val = _get(row, format_idx)
            if format_val:
                col_dict["format"] = format_val

            # Notes - convert newline-separated string to list
            notes_val = _get(row, notes_idx)
            if notes_val:
                if isinstance(notes_val, str):
                    col_dict["notes"] = [
                        line.strip() for line in notes_val.split("\n") if line.strip()
                    ]

            columns.append(col_dict)

        return columns

    def _extract_validation(
        self, workbook: openpyxl.Workbook
    ) -> tuple[dict | None, dict[str, str | None]]:
        """Extract validation rules from Validation Rules sheet.

        Columns (in order):
        - A (0): Review Notes (empty column for manual notes)
        - B (1): Column
        - C (2): Severity
        - D (3): Rule Type
        - E (4): Index
        - F (5): Description
        - G+ (6+): Dynamic key | value | key | value ... pairs

        The Index column tracks identity for merge/update/delete operations.
        Parameter pairs are read as: key_column | value_column | key_column | value_column ...

        Returns:
            Tuple of (validation_dict, review_notes_map)
            review_notes_map: Map of rule index to review note text

        """
        if self.constants.SHEET_VALIDATION not in workbook.sheetnames:
            return None, {}

        ws = workbook[self.constants.SHEET_VALIDATION]

        # Detect if Review Notes column exists by checking first header
        first_row_values = [cell.value for cell in ws[1]]
        has_review_notes = (
            first_row_values
            and first_row_values[0]
            and "review" in str(first_row_values[0]).lower()
            and "note" in str(first_row_values[0]).lower()
        )

        # Column indices depend on whether Review Notes exists
        if has_review_notes:
            # Original format with Review Notes
            col_review_note = 0
            col_column = 1
            col_severity = 2
            col_rule_type = 3
            col_index = 4
            col_description = 5
            col_generated_from = 6
            col_rule_id = 7
            col_lob = 8
            col_reason_unmappable = 9
            col_suggested_implementation = 10
            col_domain_type = 11
            col_kwargs_start = 12
        else:
            # Cleaned format without Review Notes
            col_review_note = None
            col_column = 0
            col_severity = 1
            col_rule_type = 2
            col_index = 3
            col_description = 4
            col_generated_from = 5
            col_rule_id = 6
            col_lob = 7
            col_reason_unmappable = 8
            col_suggested_implementation = 9
            col_domain_type = 10
            col_kwargs_start = 11

        expectations = []
        review_notes = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue

            # Review Notes (if column exists)
            review_note = None
            if col_review_note is not None and len(row) > col_review_note:
                review_note = row[col_review_note].value

            # Extract validation rule fields
            column = row[col_column].value if len(row) > col_column else None
            severity = row[col_severity].value if len(row) > col_severity else "critical"
            rule_type = row[col_rule_type].value if len(row) > col_rule_type else None
            rule_index = row[col_index].value if len(row) > col_index else None
            description = row[col_description].value if len(row) > col_description else ""
            generated_from = (
                row[col_generated_from].value if len(row) > col_generated_from else None
            )
            rule_id = row[col_rule_id].value if len(row) > col_rule_id else None
            lob = row[col_lob].value if len(row) > col_lob else None
            reason_unmappable = (
                row[col_reason_unmappable].value if len(row) > col_reason_unmappable else None
            )
            suggested_implementation = (
                row[col_suggested_implementation].value
                if len(row) > col_suggested_implementation
                else None
            )
            domain_type = row[col_domain_type].value if len(row) > col_domain_type else None

            # Rule type is required
            if not rule_type:
                continue

            # Build kwargs from alternating key/value columns
            kwargs = {}

            # Read key | value pairs from kwargs_start column onwards
            kwargs_col_idx = col_kwargs_start
            while kwargs_col_idx + 1 < len(row):
                key_cell = row[kwargs_col_idx]
                value_cell = row[kwargs_col_idx + 1]

                key = key_cell.value
                value = value_cell.value

                # Stop if we hit an empty key (end of key/value pairs)
                if key is None:
                    break

                if value is not None:
                    # Try to parse as JSON for complex types (lists, dicts)
                    if isinstance(value, str) and value.startswith(("[", "{")):
                        try:
                            kwargs[key] = json.loads(value)
                        except json.JSONDecodeError:
                            kwargs[key] = value
                    else:
                        kwargs[key] = value

                kwargs_col_idx += 2

            # Build expectation dict with index tracking in metadata
            # Use Excel values for generated_from and rule_id if present, otherwise use defaults
            meta: dict[str, Any] = {
                "description": description or "",
                "severity": severity or "critical",
                "generated_from": generated_from if generated_from else "user_input",
            }
            if rule_index is not None:
                meta["rule_index"] = rule_index
            if rule_id:
                meta["rule_id"] = rule_id
            if lob:
                # Parse LOB - can be comma-separated string or single value
                if isinstance(lob, str):
                    if "," in lob:
                        meta["lob"] = [item.strip() for item in lob.split(",")]
                    else:
                        meta["lob"] = [lob.strip()]
                else:
                    # Handle non-string LOB values
                    meta["lob"] = [str(lob)]
            if reason_unmappable:
                meta["reason_unmappable"] = reason_unmappable
            if suggested_implementation:
                meta["suggested_implementation"] = suggested_implementation
            if domain_type:
                meta["domain_type"] = domain_type

            exp = {
                "type": f"expect_{rule_type}",
                "kwargs": {
                    **({"column": column} if column and column != "-" else {}),
                    **kwargs,
                },
                "meta": meta,
            }

            expectations.append(exp)

            # Track review note by rule key if present
            # Key format: validation.{column}.{rule_type}.{index}
            if review_note and rule_index is not None:
                column_str = column if column and column != "-" else "table"
                rule_key = f"validation.{column_str}.{rule_type}.{rule_index}"
                review_notes[rule_key] = review_note

        if not expectations:
            return None, review_notes

        return {
            "expectations": expectations,
        }, review_notes

    def _extract_relationships(self, workbook: openpyxl.Workbook) -> dict | None:
        """Extract relationships from Relationships sheet.

        Reads columns: Source Column, References Table, References Column, Confidence,
        Type, Cardinality, Domain Context, Detection Method.
        """
        ws = workbook[self.constants.SHEET_RELATIONSHIPS]

        foreign_keys = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            source_col = row[0].value
            if not source_col:
                continue

            ref_table = row[1].value
            ref_column = row[2].value
            confidence = row[3].value if len(row) > 3 else None
            rel_type = row[4].value if len(row) > 4 else None
            # cardinality = row[5].value if len(row) > 5 else None  # Reserved for future use
            domain_context = row[6].value if len(row) > 6 else None
            detection_method = row[7].value if len(row) > 7 else None

            if not ref_table or not ref_column:
                continue

            fk = {
                "column": source_col,
                "references_table": ref_table,
                "references_column": ref_column,
            }

            if confidence:
                with contextlib.suppress(ValueError, TypeError):
                    if isinstance(confidence, (int, float, str)):
                        confidence_float = float(confidence)
                        fk["confidence"] = (
                            confidence_float / 100.0 if confidence_float > 1 else confidence_float
                        )

            if rel_type:
                fk["type"] = rel_type

            if domain_context:
                fk["domain_context"] = domain_context

            if detection_method:
                fk["detection_method"] = detection_method

            foreign_keys.append(fk)

        if not foreign_keys:
            return None

        return {
            "foreign_keys": foreign_keys,
        }

    def _extract_file_format(self, workbook: openpyxl.Workbook) -> dict | None:
        """Extract file format specification."""
        ws = workbook[self.constants.SHEET_FILE_FORMAT]

        file_format = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            field = row[0].value
            value = row[1].value

            if not field:
                continue

            # Skip empty values except for boolean fields
            if value is None or (isinstance(value, str) and not value.strip()):
                if field != "header":
                    continue

            # Type conversions
            if field == "header":
                value = bool(value)
            elif field == "skip_rows":
                value = int(value) if value and isinstance(value, (int, float, str)) else 0
            elif field == "filename_pattern":
                # Parse JSON structure for filename_pattern
                if value and isinstance(value, str):
                    try:
                        pattern_data = json.loads(value)
                        if "regex" in pattern_data and "captures" in pattern_data:
                            # Convert string keys to int keys for captures
                            captures = {
                                int(k) if isinstance(k, str) and k.isdigit() else k: v
                                for k, v in pattern_data["captures"].items()
                            }
                            value = {
                                "regex": pattern_data["regex"],
                                "captures": captures,
                            }
                    except (json.JSONDecodeError, ValueError, KeyError):
                        # Skip invalid JSON
                        continue

            file_format[field] = value

        return file_format if file_format else None

    def _extract_metadata(self, workbook: openpyxl.Workbook) -> dict | None:
        """Extract metadata from Metadata sheet."""
        if self.constants.SHEET_METADATA not in workbook.sheetnames:
            return None

        ws = workbook[self.constants.SHEET_METADATA]

        metadata = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            field = row[0].value
            value = row[1].value

            if not field:
                continue

            # Skip empty values
            if value is None or (isinstance(value, str) and not value.strip()):
                continue

            # Type conversions
            if field == "pipeline_phase":
                value = int(value) if value and isinstance(value, (int, float, str)) else None

            metadata[field] = value

        return metadata if metadata else None
